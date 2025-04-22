#!/usr/bin/env python
from flask import Flask, request, send_file, render_template_string
import io
from datetime import datetime, timedelta
import pytz
from icalendar import Calendar, Event
from openpyxl import load_workbook

app = Flask(__name__)

# HTML form template with Dracula theme styles
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Study Schedule Generator</title>
    <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
    <style>
        :root {
            --dracula-background: #282a36;
            --dracula-current-line: #44475a;
            --dracula-foreground: #f8f8f2;
            --dracula-comment: #6272a4;
            --dracula-cyan: #8be9fd;
            --dracula-green: #50fa7b;
            --dracula-orange: #ffb86c;
            --dracula-pink: #ff79c6;
            --dracula-purple: #bd93f9;
            --dracula-red: #ff5555;
            --dracula-yellow: #f1fa8c;
        }

        body {
            background-color: var(--dracula-background);
            color: var(--dracula-foreground);
        }

        .container {
            background-color: var(--dracula-current-line);
            border-radius: 10px;
            padding: 30px;
            margin-top: 30px;
            margin-bottom: 30px;
        }

        .form-control {
            background-color: var(--dracula-background) !important;
            border-color: var(--dracula-purple) !important;
            color: var(--dracula-foreground) !important;
        }

        .form-control:focus {
            box-shadow: 0 0 0 0.2rem rgba(189, 147, 249, 0.25) !important;
        }

        .btn-primary {
            background-color: var(--dracula-purple) !important;
            border-color: var(--dracula-purple) !important;
        }

        .btn-primary:hover {
            background-color: var(--dracula-pink) !important;
            border-color: var(--dracula-pink) !important;
        }

        .btn-info {
            background-color: var(--dracula-cyan) !important;
            border-color: var(--dracula-cyan) !important;
            color: var(--dracula-background) !important;
        }

        .btn-info:hover {
            background-color: var(--dracula-green) !important;
            border-color: var(--dracula-green) !important;
        }

        label {
            color: var(--dracula-cyan);
            font-weight: bold;
        }

        h2 {
            color: var(--dracula-pink);
            margin-bottom: 30px;
        }

        select option {
            background-color: var(--dracula-background);
            color: var(--dracula-foreground);
        }
    </style>
    <script>
    function toggleInput() {
        var listType = document.getElementById("list_type").value;
        document.getElementById("textInputBox").style.display = listType === "udemy" ? "block" : "none";
        document.getElementById("fileInputBox").style.display = listType === "spreadsheet" ? "block" : "none";
    }
    </script>
</head>
<body>
    <div class="container mt-5">
        <h2>Generate Study Schedule</h2>
        <form method="post" enctype="multipart/form-data">
            <div class="form-group">
                <label>Start Date (YYYY-MM-DD):</label>
                <input type="text" class="form-control" name="start_date" required>
            </div>
            <div class="form-group">
                <label>Study Days (0=Monday, ..., 6=Sunday), separated by commas:</label>
                <input type="text" class="form-control" name="study_days" required>
            </div>
            <div class="form-group">
                <label>Start Time (HH:MM):</label>
                <input type="text" class="form-control" name="start_time" required>
            </div>
            <div class="form-group">
                <label>Daily Study Limit Hours:</label>
                <input type="number" class="form-control" name="daily_study_limit_hours" required>
            </div>
            <div class="form-group">
                <label>Multiplier:</label>
                <input type="text" class="form-control" name="multiplier" required>
            </div>
            <div class="form-group">
                <label>Input Type:</label>
                <select class="form-control" name="list_type" id="list_type" onchange="toggleInput()" required>
                    <option value="udemy">Udemy List</option>
                    <option value="spreadsheet">Spreadsheet</option>
                </select>
            </div>
            <div id="textInputBox" class="form-group">
                <label>Udemy Class Input:</label>
                <textarea class="form-control" name="class_input" rows="10"></textarea>
            </div>
            <div id="fileInputBox" class="form-group" style="display:none;">
                <label>Upload Spreadsheet:</label>
                <input type="file" class="form-control-file" name="spreadsheet" accept=".xlsx">
            </div>
            <input type="submit" class="btn btn-primary" value="Generate Schedule">
        </form>
    </div>
    <script src="https://code.jquery.com/jquery-3.5.1.slim.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/@popperjs/core@2.5.2/dist/umd/popper.min.js"></script>
    <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
</body>
</html>
"""

@app.route('/', methods=['GET'])
def form():
    return render_template_string(HTML_TEMPLATE)

@app.route('/', methods=['POST'])
def generate_schedule():
    start_date_str = request.form['start_date']
    study_days_input = request.form['study_days']
    start_time_str = request.form['start_time']
    daily_study_limit_hours = int(request.form['daily_study_limit_hours'])
    multiplier = float(request.form['multiplier'])
    list_type = request.form['list_type']

    if list_type == 'udemy':
        classes = parse_class_input(request.form['class_input'])
    else:  # spreadsheet
        if 'spreadsheet' not in request.files:
            return 'No file uploaded', 400
        file = request.files['spreadsheet']
        classes = parse_spreadsheet(file)

    adjusted_classes = apply_multiplier(classes, multiplier)

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    study_days = [int(day) for day in study_days_input.split(',')]
    study_schedule = schedule_classes(adjusted_classes, start_date, study_days, daily_study_limit_hours)

    ical_data = create_calendar_events(study_schedule, start_time_str, 'America/Sao_Paulo', daily_study_limit_hours)

    ical_file = io.BytesIO(ical_data)
    ical_file.seek(0)
    return send_file(ical_file, as_attachment=True, download_name='study_schedule.ics', mimetype='text/calendar')

def parse_class_input(class_input):
    classes = []
    lines = class_input.split('\n')

    i = 0
    while i < len(lines):
        status = lines[i].strip()
        if i + 2 < len(lines):
            subject = lines[i + 1].strip()
            duration_line = lines[i + 2].strip()
            if 'min' in duration_line:
                duration_str = duration_line.replace('min', '')
                try:
                    duration = int(duration_str)
                    classes.append([status, subject, duration])
                    i += 3
                except ValueError:
                    i += 1
            else:
                i += 1
        else:
            break
    return classes

def parse_spreadsheet(file_storage):
    """Parses the uploaded Excel spreadsheet and returns a list of classes."""
    classes = []
    wb = load_workbook(filename=io.BytesIO(file_storage.read()))
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2):  # Skip header row
        if not row[0].value:  # Skip empty rows
            continue
            
        day = row[0].value
        title = row[1].value
        duration = row[3].value  # This is now a datetime.time object
        completed = row[4].value == 'x' if row[4].value else False
        
        # Skip completed classes
        if completed:
            continue
            
        try:
            # Handle duration as time object
            if hasattr(duration, 'hour') and hasattr(duration, 'minute'):
                # Convert time object to minutes
                duration_minutes = duration.hour * 60 + duration.minute
                if duration.second > 0:
                    duration_minutes += 1  # Round up if there are seconds
            elif isinstance(duration, str) and ':' in duration:
                # Fallback for string format
                h, m, s = map(int, duration.split(':'))
                duration_minutes = h * 60 + m + (1 if s > 0 else 0)
            else:
                continue  # Skip if no valid duration
                
            # Format the title with the day
            subject = f"🚗 {day}: {title}" if "Pista Rápida" in title else f"{day}: {title}"
            classes.append(["Not Started", subject, duration_minutes])
        except (ValueError, AttributeError):
            continue  # Skip invalid duration formats
    
    return classes

def apply_multiplier(classes, multiplier):
    return [[cls[0], cls[1], cls[2] * multiplier] for cls in classes]

def schedule_classes(classes, start_date, study_days, daily_study_limit_hours):
    study_schedule = {}
    daily_study_limit_minutes = daily_study_limit_hours * 60
    current_date = start_date
    time_left = daily_study_limit_minutes

    for cls in classes:
        while cls[2] > 0:
            if current_date.weekday() in study_days:
                if time_left >= cls[2]:
                    if current_date not in study_schedule:
                        study_schedule[current_date] = []
                    study_schedule[current_date].append(cls)
                    time_left -= cls[2]
                    cls[2] = 0
                else:
                    if current_date not in study_schedule:
                        study_schedule[current_date] = []
                    split_class = cls[:]
                    split_class[2] = time_left
                    study_schedule[current_date].append(split_class)
                    cls[2] -= time_left
                    time_left = 0
            if time_left == 0 or current_date.weekday() not in study_days:
                current_date += timedelta(days=1)
                time_left = daily_study_limit_minutes

    return study_schedule

def create_calendar_events(study_schedule, start_time_str, timezone, daily_study_limit_hours):
    cal = Calendar()
    cal.add('prodid', '-//Study Schedule Calendar//mxm.dk//')
    cal.add('version', '2.0')
    local_tz = pytz.timezone(timezone)

    for day, classes in study_schedule.items():
        start_datetime = datetime.combine(day, datetime.strptime(start_time_str, '%H:%M').time())
        start_datetime = local_tz.localize(start_datetime)
        end_datetime = start_datetime + timedelta(minutes=daily_study_limit_hours*60)

        event = Event()
        event.add('summary', f"Study Block {day.strftime('%Y-%m-%d')}")
        event.add('dtstart', start_datetime)
        event.add('dtend', end_datetime)
        event.add('description', "Classes today:\n" + '\n'.join(cls[1] for cls in classes))
        cal.add_component(event)

    return cal.to_ical()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
