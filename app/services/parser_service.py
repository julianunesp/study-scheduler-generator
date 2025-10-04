"""Service for parsing different input formats (Udemy list and Excel spreadsheet)."""
import io
from openpyxl import load_workbook


def parse_udemy_list(class_input):
    """
    Parse Udemy class list from text input.
    
    Args:
        class_input (str): Text input containing classes in format:
            status
            subject
            duration in minutes
    
    Returns:
        list: List of [status, subject, duration] for each class
    """
    classes = []
    lines = class_input.split('\n')

    i = 0
    while i < len(lines):
        status = lines[i].strip()
        if i + 2 < len(lines):
            subject = lines[i + 1].strip()
            duration_line = lines[i + 2].strip()
            if 'min' in duration_line:
                duration_str = duration_line.replace('min', '')
                try:
                    duration = int(duration_str)
                    classes.append([status, subject, duration])
                    i += 3
                except ValueError:
                    i += 1
            else:
                i += 1
        else:
            break
    return classes


def parse_spreadsheet(file_storage):
    """
    Parse Excel spreadsheet with class information.
    
    Args:
        file_storage: FileStorage object from Flask request
    
    Returns:
        list: List of [status, subject, duration] for each class
    """
    classes = []
    wb = load_workbook(filename=io.BytesIO(file_storage.read()))
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2):  # Skip header row
        if not row[0].value:  # Skip empty rows
            continue
            
        day = row[0].value
        title = row[1].value
        duration = row[3].value  # This is now a datetime.time object
        completed = row[4].value == 'x' if row[4].value else False
        
        # Skip completed classes
        if completed:
            continue
            
        try:
            # Handle duration as time object
            if hasattr(duration, 'hour') and hasattr(duration, 'minute'):
                # Convert time object to minutes
                duration_minutes = duration.hour * 60 + duration.minute
                if duration.second > 0:
                    duration_minutes += 1  # Round up if there are seconds
            elif isinstance(duration, str) and ':' in duration:
                # Fallback for string format
                h, m, s = map(int, duration.split(':'))
                duration_minutes = h * 60 + m + (1 if s > 0 else 0)
            else:
                continue  # Skip if no valid duration
                
            # Format the title with the day
            subject = f"🚗 {day}: {title}" if "Pista Rápida" in title else f"{day}: {title}"
            classes.append(["Not Started", subject, duration_minutes])
        except (ValueError, AttributeError):
            continue  # Skip invalid duration formats
    
    return classes

