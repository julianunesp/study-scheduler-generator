"""Routes for Study Schedule Generator."""
import os
import io
import base64
from datetime import datetime
from flask import Blueprint, request, send_file, render_template, jsonify
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .services import (
    parse_udemy_list,
    parse_spreadsheet,
    apply_multiplier,
    schedule_classes,
    create_calendar_events,
    get_html_parser,
    get_google_service
)
from .config import ALLOWED_EXTENSIONS, MAX_FILE_SIZES

main_bp = Blueprint('main', __name__)


def get_file_extension(filename):
    """Extract file extension from filename."""
    if '.' in filename:
        return filename.rsplit('.', 1)[1].lower()
    return None


def validate_file_upload(file, expected_extensions):
    """
    Validate uploaded file for type and size restrictions.
    
    Args:
        file: FileStorage object from request.files
        expected_extensions: List of allowed extensions for this upload
        
    Returns:
        tuple: (is_valid, error_message)
    """
    if not file or file.filename == '':
        return False, 'No file selected'
    
    # Secure the filename
    filename = secure_filename(file.filename)
    
    # Check file extension
    file_ext = get_file_extension(filename)
    if not file_ext or file_ext not in expected_extensions:
        allowed = ', '.join(expected_extensions)
        return False, f'Invalid file type. Allowed types: {allowed}'
    
    # Check if extension is in our allowed list
    if file_ext not in ALLOWED_EXTENSIONS:
        return False, f'File type .{file_ext} is not allowed'
    
    # Check MIME type
    if file.content_type not in ALLOWED_EXTENSIONS[file_ext]:
        return False, f'Invalid content type for .{file_ext} file'
    
    # Check file size
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)  # Reset file pointer
    
    max_size = MAX_FILE_SIZES.get(file_ext, 0)
    if file_size > max_size:
        max_size_mb = max_size / (1024 * 1024)
        return False, f'File size exceeds maximum allowed ({max_size_mb:.1f}MB)'
    
    if file_size == 0:
        return False, 'File is empty'
    
    return True, None


def create_classes_spreadsheet(classes):
    """
    Create an Excel spreadsheet with the extracted classes.
    
    Args:
        classes (list): List of [status, subject, duration] for each class
        
    Returns:
        bytes: Excel file data in bytes format
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Course Classes"
    
    # Header styling
    header_fill = PatternFill(start_color="BD93F9", end_color="BD93F9", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    headers = ["#", "Module/Class", "Duration (HH:MM:SS)", "Status", "Notes"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data rows
    for idx, cls in enumerate(classes, 1):
        status, subject, duration_minutes = cls
        
        # Convert duration to HH:MM:SS
        total_seconds = int(duration_minutes * 60)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        
        ws.cell(row=idx+1, column=1).value = idx
        ws.cell(row=idx+1, column=2).value = subject
        ws.cell(row=idx+1, column=3).value = duration_str
        ws.cell(row=idx+1, column=4).value = "⬜ Not Started"
        ws.cell(row=idx+1, column=5).value = ""
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 30
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.read()


@main_bp.route('/', methods=['GET'])
def index():
    """Render the main form page."""
    return render_template('index.html')


@main_bp.route('/download-sample')
def download_sample():
    """Serve the sample spreadsheet for download."""
    sample_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'sample_spreadsheet.xlsx')
    if os.path.exists(sample_path):
        return send_file(
            sample_path,
            as_attachment=True,
            download_name='sample_spreadsheet.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        return 'Sample file not found', 404


@main_bp.route('/generate', methods=['POST'])
def generate_schedule():
    """Process form data and generate study schedule."""
    try:
        # Extract form data
        start_date_str = request.form['start_date']
        study_days_list = request.form.getlist('study_days')
        start_time_str = request.form['start_time']
        daily_study_limit_hours = int(request.form['daily_study_limit_hours'])
        multiplier = float(request.form['multiplier'])
        list_type = request.form['list_type']
        course_name = request.form.get('course_name', 'Study')

        # Parse classes based on input type
        if list_type == 'udemy':
            classes = parse_udemy_list(request.form['class_input'])
        elif list_type == 'html':
            # Handle HTML upload with AI parsing
            if 'html_file' not in request.files:
                return jsonify({'success': False, 'error': 'No HTML file uploaded'}), 400
            html_file = request.files['html_file']
            
            # Validate HTML file
            is_valid, error_msg = validate_file_upload(html_file, ['html'])
            if not is_valid:
                return jsonify({'success': False, 'error': error_msg}), 400
            
            html_content = html_file.read().decode('utf-8')
            
            try:
                # Use AI agent to parse HTML
                parser = get_html_parser()
                classes = parser.parse_html_to_classes_list(html_content)
                
                if not classes:
                    return jsonify({'success': False, 'error': 'Could not extract course content from HTML'}), 400
            except Exception as e:
                return jsonify({'success': False, 'error': f'Error parsing HTML: {str(e)}'}), 500
        else:  # spreadsheet
            if 'spreadsheet' not in request.files:
                return jsonify({'success': False, 'error': 'No file uploaded'}), 400
            file = request.files['spreadsheet']
            
            # Validate spreadsheet file
            is_valid, error_msg = validate_file_upload(file, ['xlsx', 'xls'])
            if not is_valid:
                return jsonify({'success': False, 'error': error_msg}), 400
            
            classes = parse_spreadsheet(file)

        # Apply time multiplier
        adjusted_classes = apply_multiplier(classes, multiplier)

        # Schedule classes
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        study_days = [int(day) for day in study_days_list]
        study_schedule = schedule_classes(
            adjusted_classes,
            start_date,
            study_days,
            daily_study_limit_hours
        )

        # Create iCalendar file
        ical_data = create_calendar_events(
            study_schedule,
            start_time_str,
            'America/Sao_Paulo',
            daily_study_limit_hours,
            course_name
        )
        
        # Create spreadsheet with extracted classes
        spreadsheet_data = create_classes_spreadsheet(classes)

        # Return JSON with base64-encoded data
        return jsonify({
            'success': True,
            'calendar': base64.b64encode(ical_data).decode('utf-8'),
            'spreadsheet': base64.b64encode(spreadsheet_data).decode('utf-8')
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/google/auth')
def google_auth():
    """Initiate Google OAuth flow."""
    try:
        google_service = get_google_service()
        
        # Generate state for security
        import secrets
        state = secrets.token_urlsafe(32)
        
        # Store state in session (you might want to use Flask session here)
        authorization_url, _ = google_service.get_authorization_url(state)
        
        return jsonify({
            'success': True,
            'authorization_url': authorization_url,
            'state': state
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/oauth2callback')
def oauth2callback():
    """Handle OAuth callback from Google."""
    try:
        google_service = get_google_service()
        
        # Get the full callback URL
        authorization_response = request.url
        
        # Exchange code for credentials
        google_service.handle_oauth_callback(authorization_response)
        
        # Redirect back to app with success message
        return render_template('oauth_success.html')
    except Exception as e:
        return f'Authentication failed: {str(e)}', 400


@main_bp.route('/google/calendars')
def list_google_calendars():
    """List user's Google Calendars."""
    try:
        google_service = get_google_service()
        
        # Try to load existing credentials
        if not google_service.load_credentials():
            return jsonify({
                'success': False,
                'error': 'Not authenticated',
                'needs_auth': True
            }), 401
        
        calendars = google_service.list_calendars()
        
        return jsonify({
            'success': True,
            'calendars': [
                {
                    'id': cal['id'],
                    'summary': cal['summary'],
                    'primary': cal.get('primary', False)
                }
                for cal in calendars
            ]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/google/import-calendar', methods=['POST'])
def import_to_google_calendar():
    """Import iCal data to Google Calendar."""
    try:
        data = request.get_json()
        ical_data_b64 = data.get('calendar_data')
        calendar_id = data.get('calendar_id', 'primary')
        
        if not ical_data_b64:
            return jsonify({'success': False, 'error': 'No calendar data provided'}), 400
        
        # Decode base64 data
        ical_data = base64.b64decode(ical_data_b64)
        
        google_service = get_google_service()
        
        # Try to load existing credentials
        if not google_service.load_credentials():
            return jsonify({
                'success': False,
                'error': 'Not authenticated',
                'needs_auth': True
            }), 401
        
        # Import events
        event_ids = google_service.import_ical_to_calendar(ical_data, calendar_id)
        
        return jsonify({
            'success': True,
            'message': f'Successfully imported {len(event_ids)} events',
            'event_count': len(event_ids)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main_bp.route('/google/upload-drive', methods=['POST'])
def upload_to_google_drive():
    """Import Excel spreadsheet as Google Sheets."""
    try:
        data = request.get_json()
        spreadsheet_data_b64 = data.get('spreadsheet_data')
        filename = data.get('filename', 'Course Classes')
        
        if not spreadsheet_data_b64:
            return jsonify({'success': False, 'error': 'No spreadsheet data provided'}), 400
        
        # Decode base64 data
        spreadsheet_data = base64.b64decode(spreadsheet_data_b64)
        
        google_service = get_google_service()
        
        # Try to load existing credentials
        if not google_service.load_credentials():
            return jsonify({
                'success': False,
                'error': 'Not authenticated',
                'needs_auth': True
            }), 401
        
        # Import Excel as Google Sheets
        file_info = google_service.import_excel_to_sheets(
            spreadsheet_data,
            filename
        )
        
        return jsonify({
            'success': True,
            'message': 'Spreadsheet imported successfully to Google Sheets',
            'file': {
                'id': file_info['id'],
                'name': file_info['name'],
                'webViewLink': file_info.get('webViewLink'),
                'mimeType': file_info.get('mimeType')
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

